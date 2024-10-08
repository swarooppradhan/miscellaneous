import pandas as pd
import trino
from trino.auth import BasicAuthentication
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import getpass
import logging
import datetime
import sqlparse
import re
import os
import time
import threading
import argparse

# Global flag to indicate when the execution is complete
execution_complete = False
# Dictionary to store reusable connections
connection_pool = {}
# Dictionary to store variable values
variable_values_cache = {}
# Dictionary to keep track of failed connections
failed_connections = set()
# A lock to ensure thread-safe updates
data_lock = threading.Lock()

def main():
    # Set up argument parsing
    parser = argparse.ArgumentParser(
        description="Script to execute SQL test cases for Trino based on an Excel file input. "
                    "It runs Setup, Test, and Clean up cases based on selected teams and environment."
    )
    parser.add_argument('-f', '--file', type=str, help='Path to the Excel file')
    parser.add_argument('-e', '--env', type=str, help='Environment name')
    parser.add_argument('-t', '--teams', type=str, help='Comma-separated team names or "All Teams"')
    parser.add_argument('-r', '--refresh', type=int, help='Refresh frequency in minutes for the summary display')

    # Parse the arguments
    args = parser.parse_args()

    # Check if the Excel file path is provided, otherwise prompt for it
    if args.file:
        file_path = os.path.abspath(args.file)
    else:
        file_path = os.path.abspath(input("Enter the path to the Excel file: ").strip())

    # Read the Excel sheets using sheet names
    test_cases_df = pd.read_excel(file_path, sheet_name='Test Cases').fillna('')
    users_df = pd.read_excel(file_path, sheet_name='Users').fillna('')
    trino_env_df = pd.read_excel(file_path, sheet_name='Trino Env').fillna('')
    sql_variables_df = pd.read_excel(file_path, sheet_name='SQL Variables').fillna('')

    # Check if the environment is provided, otherwise prompt for it
    if args.env:
        selected_env = args.env
    else:
        selected_env = get_selected_env(trino_env_df)

    # Check if teams are provided, otherwise prompt for it
    if args.teams:
        if args.teams.lower() == "all teams":
            selected_teams = sorted(test_cases_df['Team'].unique())
        else:
            selected_teams = [team.strip() for team in args.teams.split(',')]
    else:
        selected_teams = get_selected_teams(test_cases_df)

    # Check if refresh frequency is provided, otherwise prompt for it
    if args.refresh:
        refresh_frequency = args.refresh
    else:
        refresh_frequency = int(input("Enter the refresh frequency in minutes for the summary display: ").strip())

    # Generate log file and result file names without team names
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    input_file_name = os.path.splitext(os.path.basename(file_path))[0]
    excel_directory = os.path.dirname(file_path)
    result_filepath = os.path.join(excel_directory, f"{input_file_name}_results_{selected_env}_{timestamp}.xlsx").replace(" ", "_")
    main_log_filepath = os.path.join(excel_directory, f"{input_file_name}_main_log_{selected_env}_{timestamp}.log").replace(" ", "_")
    
    setup_logging(main_log_filepath, "main", to_console=True)

    # Get the main logger
    main_logger = logging.getLogger("main")

    # Log the selections using the main logger
    main_logger.info(f"Excel File Path: {file_path}")
    main_logger.info(f"Selected Environment: {selected_env}")
    main_logger.info(f"Selected Teams: {', '.join(selected_teams)}")
    main_logger.info(f"Refresh Frequency (minutes): {refresh_frequency}")

    # Filter test cases based on Execution Type
    setup_test_cases = test_cases_df[test_cases_df['Execution Type'] == 'Setup']
    cleanup_test_cases = test_cases_df[test_cases_df['Execution Type'] == 'Clean up']
    
    # Select only the 'Test' test cases for the selected teams
    team_test_cases_df = test_cases_df[
        (test_cases_df['Execution Type'] == 'Test') & 
        (test_cases_df['Team'].isin(selected_teams))
    ]

    # Combine all test cases in the correct order: Setup -> Test -> Clean up
    ordered_test_cases_df = pd.concat([setup_test_cases, team_test_cases_df, cleanup_test_cases], ignore_index=True)
    
    # Set up loggers for all teams present in the test cases for which test cases will be executed
    all_teams = sorted(ordered_test_cases_df['Team'].unique())
    
    main_logger.info("Setting up log files for all teams...")
    
    log_file_paths = {}
    for team in all_teams:
        log_filepath = os.path.join(excel_directory, f"{input_file_name}_log_{selected_env}_{team}_{timestamp}.log").replace(" ", "_")
        setup_logging(log_filepath, team)
        log_file_paths[team] = log_filepath
        main_logger.info(f"Log file for team '{team}': {log_filepath}")

    main_logger.info("Log setup completed. Refer to the above paths for team-specific logs.")

    # Filter users for the selected environment and required groups
    users_df = users_df[
        (users_df['Env'] == selected_env) &
        (users_df['Group'].isin(ordered_test_cases_df['Group'].unique()))
    ]

    # Log users for the selected environment
    main_logger.info("Logging users for the selected environment:")
    for group in users_df['Group'].unique():
        user = users_df[users_df['Group'] == group]['User'].iloc[0]
        main_logger.info(f"Group: {group}, User: {user}")

    # Handle password retrieval using group names
    user_passwords = get_user_passwords(users_df)

    process_test_cases(result_filepath, selected_env, selected_teams, user_passwords, ordered_test_cases_df, users_df, trino_env_df, sql_variables_df, refresh_frequency)

    # After test execution, log the result file path and all log file paths
    main_logger.info(f"Test execution completed. Results have been saved to: {result_filepath}")
    main_logger.info("The following team-specific log files were generated:")
    for team, path in log_file_paths.items():
        if os.path.exists(path) and os.path.getsize(path) > 0:
            main_logger.info(f"Team '{team}': {path}")
        else:
            main_logger.warn(f"No log entries were made for team '{team}', so the log file may be empty or not created.")

def save_filtered_sheets(result_filepath, selected_env, users_df, trino_env_df, sql_variables_df):
    # Save the filtered Users sheet
    filtered_users_df = users_df[users_df['Env'] == selected_env]
    with pd.ExcelWriter(result_filepath, engine='openpyxl', mode='a') as writer:
        filtered_users_df.to_excel(writer, sheet_name='Users', index=False)

    # Save the filtered Trino Env sheet
    filtered_trino_env_df = trino_env_df[trino_env_df['Env'] == selected_env]
    with pd.ExcelWriter(result_filepath, engine='openpyxl', mode='a') as writer:
        filtered_trino_env_df.to_excel(writer, sheet_name='Trino Env', index=False)

    # Save the filtered SQL Variables sheet
    filtered_sql_variables_df = sql_variables_df[sql_variables_df['Env'] == selected_env]
    with pd.ExcelWriter(result_filepath, engine='openpyxl', mode='a') as writer:
        filtered_sql_variables_df.to_excel(writer, sheet_name='SQL Variables', index=False)

# Function to display the execution summary
def display_summary(ordered_test_cases_df, total_test_cases, refresh_frequency):
    # Get the main logger
    main_logger = logging.getLogger("main")

    while not execution_complete:
        with data_lock:
            executed_cases = ordered_test_cases_df[ordered_test_cases_df['Actual Status'].notna()].shape[0]
            passed_cases = ordered_test_cases_df[ordered_test_cases_df['Result'] == 'PASS'].shape[0]
            failed_cases = ordered_test_cases_df[ordered_test_cases_df['Result'] == 'FAIL'].shape[0]

        main_logger.info("=" * 50)
        main_logger.info(f"Total Test Cases: {total_test_cases}")
        main_logger.info(f"Executed Test Cases: {executed_cases}")
        main_logger.info(f"Passed Test Cases: {passed_cases}")
        main_logger.info(f"Failed Test Cases: {failed_cases}")
        main_logger.info("=" * 50)
        
        time.sleep(refresh_frequency * 60)  # Refresh based on user-defined frequency

    # Final summary once execution is complete
    with data_lock:
        executed_cases = ordered_test_cases_df[ordered_test_cases_df['Actual Status'].notna()].shape[0]
        passed_cases = ordered_test_cases_df[ordered_test_cases_df['Result'] == 'PASS'].shape[0]
        failed_cases = ordered_test_cases_df[ordered_test_cases_df['Result'] == 'FAIL'].shape[0]

    main_logger.info("Final Summary")
    main_logger.info("=" * 50)
    main_logger.info(f"Total Test Cases: {total_test_cases}")
    main_logger.info(f"Executed Test Cases: {executed_cases}")
    main_logger.info(f"Passed Test Cases: {passed_cases}")
    main_logger.info(f"Failed Test Cases: {failed_cases}")
    main_logger.info("=" * 50)
    
# Function to set up logging with the desired log file name
def setup_logging(log_filepath, logger_name,  to_console=False):
    # Set up logging for the given logger name
    logger = logging.getLogger(logger_name)
    logger.setLevel(logging.INFO)

    # Create a file handler for the log file
    file_handler = logging.FileHandler(log_filepath, encoding='utf-8')
    file_handler.setLevel(logging.INFO)

    # Create a logging format
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)

    # Add the file handler to the logger
    if logger.hasHandlers():
        logger.handlers.clear()  # Remove existing handlers to prevent duplicate logs
    logger.addHandler(file_handler)
    
    if to_console:
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
    
    # Prevent the logger from propagating messages to the root logger
    logger.propagate = False

# Function to get passwords for unique users
def get_user_passwords(users_df):
    # Filter users by the selected environment
    env_users_groups = users_df[['User', 'Group']].drop_duplicates()

    user_passwords = {}
    
    for _, row in env_users_groups.iterrows():
        user = row['User']
        group = row['Group'].upper()  # Convert group to uppercase
        env_var_name = f"{group}_SECRET"
        
        if env_var_name in os.environ:
            user_passwords[user] = os.environ[env_var_name]
        else:
            # Use getpass to hide password input
            user_passwords[user] = getpass.getpass(prompt=f"Enter the password for group {group} (user: {user}): ")
    
    return user_passwords

# Function to get or reuse a Trino connection using BasicAuthentication
def get_or_create_trino_connection(host_url, user, password):
    connection_key = (host_url, user)
    
    if connection_key not in connection_pool:
        # Create a new connection if it doesn't exist in the pool
        connection_pool[connection_key] = trino.dbapi.connect(
            host=host_url,
            port=443,  # Using HTTPS port 443
            user=user,
            auth=BasicAuthentication(user, password),  # Use BasicAuthentication
            http_scheme='https'  # Use 'https' for secure connections
        )
        
    return connection_pool[connection_key]

# Function to pretty format the SQL query
def format_sql(sql_query):
    return sqlparse.format(sql_query, reindent=True, keyword_case='upper')

# Function to collect variable values from the "SQL Variables" sheet and prompt if not found
def collect_variable_values(test_cases_df, sql_variables_df, selected_env):
    # Extract all unique variables from all SQL queries in the test cases
    all_sql_queries = ' '.join(test_cases_df['SQL Query'].dropna().tolist())
    variables = re.findall(r"##(.*?)##", all_sql_queries)
    
    # Get values for the selected environment from the SQL Variables sheet
    env_variable_values = sql_variables_df[sql_variables_df['Env'] == selected_env]

    # Prompt for each unique variable only once if not found in the SQL Variables sheet
    unique_variables = set(variables)
    for var in unique_variables:
        # Check if the variable exists in the SQL Variables sheet for the selected environment
        var_value_row = env_variable_values[env_variable_values['Variable'] == var]
        
        if not var_value_row.empty:
            value = var_value_row.iloc[0]['Value']
            variable_values_cache[var] = value  # Store the value in the cache
        elif var not in variable_values_cache:
            # Prompt user to enter a value for each variable if not already cached
            value = input(f"Enter value for variable '{var}': ")
            variable_values_cache[var] = value
            
# Function to replace variables in a given SQL query using collected values and remove any trailing semicolon
def replace_variables_in_sql(sql_query):
    for var, value in variable_values_cache.items():
        sql_query = sql_query.replace(f"##{var}##", value)
    
    # Remove any trailing semicolon from the query
    sql_query = sql_query.rstrip(';').strip()
    return sql_query

# Function to execute SQL and return status
def execute_sql_with_trino(conn, sql_query, logger):
    try:
        cursor = conn.cursor()
        cursor.execute(sql_query)
        
        if sql_query.strip().lower().startswith(('select', 'with', 'show')):
            results = cursor.fetchall()
            logger.info(f"Query executed successfully: \n{format_sql(sql_query)}")
            return "COMPLETED", results
        else:
            logger.info(f"DDL executed successfully: \n{format_sql(sql_query)}")
            return "COMPLETED", "DDL statement executed"

    except Exception as e:
        logger.error(f"Error executing query: \n{format_sql(sql_query)}. Error: {str(e)}")
        return "ERROR", str(e)

# Function to generate the output filename
def generate_output_filename(excel_directory, team, env, timestamp):
    result_filename = f"trino_test_results_{team or 'All'}_{env}_{timestamp}.xlsx".replace(' ', '_')
    return os.path.join(excel_directory, result_filename)

# Function to save results to a new Excel file
def save_results_to_new_excel(file_path, df, sheet_name='Test Cases'):
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# Function to apply formatting (PASS = Green, FAIL = Red)
def apply_result_formatting(file_path, df, sheet_name='Test Cases'):
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    for index, row in df.iterrows():
        result_cell = sheet[f'L{index + 2}']  # Column L is the 'Result'
        if row['Result'] == 'PASS':
            result_cell.fill = green_fill
        elif row['Result'] == 'FAIL':
            result_cell.fill = red_fill

    workbook.save(file_path)

# Function to get team selection
def get_selected_teams(test_cases_df):
    # Get a unique list of teams and present them as a numbered list
    teams = sorted(test_cases_df['Team'].unique())
    print("Available teams:")
    for idx, team in enumerate(teams, 1):
        print(f"{idx}. {team}")

    print("0. All Teams")  # Option to select all teams

    # Prompt user to enter team numbers as a comma-separated list
    selected_team_numbers = input("Enter the team numbers to execute (comma-separated, e.g., 1,3,5) or '0' for all: ")

    # Convert the input into a list of selected team indices
    selected_indices = [int(num.strip()) for num in selected_team_numbers.split(",") if num.strip().isdigit()]

    # Determine the actual selected teams
    if 0 in selected_indices:
        return teams  # Return all teams if '0' is selected
    else:
        return [teams[idx - 1] for idx in selected_indices if 1 <= idx <= len(teams)]

# Function to get environment selection
def get_selected_env(trino_env_df):
    envs = sorted(trino_env_df['Env'].unique())
    
    print("\nSelect the environment to be used:")
    for idx, env in enumerate(envs, start=1):
        print(f"{idx}. {env}")

    selected_env_option = int(input("Enter the number corresponding to your choice: "))
    selected_env = envs[selected_env_option - 1]
    
    return selected_env

def execute_test_cases(ordered_test_cases_df, trino_env_df, users_df, selected_env, user_passwords, sql_variables_df, execution_type, team=None):
    current_thread = threading.current_thread().name  # Get the current thread's name

    for index, test_case in ordered_test_cases_df.iterrows():
        # Skip test cases that do not match the execution type or team
        if test_case['Execution Type'] != execution_type or (team is not None and test_case['Team'] != team):
            continue
        
        test_case_number = test_case['Test Case Number']
        team = test_case['Team']
        instance_type = test_case['Trino Instance Type']
        sql_query = test_case['SQL Query']
        expected_status = test_case['Expected Status']
        expected_response = test_case['Expected Response'].strip()
        group = test_case['Group']
        use_case = test_case['Use Case']

        # Obtain the logger for the specific team
        logger = logging.getLogger(team)

        logger.info(f"[{current_thread}] Executing {execution_type} Test Case Number: {test_case_number} | Team: {team}")
        logger.info(f"[{current_thread}] Use Case: {use_case}")

        # Replace variables in the SQL query using collected values
        executed_sql = replace_variables_in_sql(sql_query)
        logger.info(f"[{current_thread}] SQL query after variable replacement: \n{format_sql(executed_sql)}")

        # Update the 'Executed SQL' column directly in the original DataFrame using .loc
        with data_lock:
            ordered_test_cases_df.loc[index, 'Executed SQL'] = executed_sql

        trino_env_row = trino_env_df[
            (trino_env_df['Team'] == team) &
            (trino_env_df['Trino Instance Type'] == instance_type) &
            (trino_env_df['Env'] == selected_env)
        ]

        if trino_env_row.empty:
            with data_lock:
                ordered_test_cases_df.loc[index, 'Actual Status'] = 'ERROR'
                ordered_test_cases_df.loc[index, 'Result'] = 'FAIL'
                ordered_test_cases_df.loc[index, 'Error Message'] = 'Host URL not found'
            logger.info(f"[{current_thread}] Host URL not found for Team: {team}, Instance Type: {instance_type}, and Env: {selected_env}")
            continue

        host_url = trino_env_row.iloc[0]['Host URL']
        user_row = users_df[(users_df['Env'] == selected_env) & (users_df['Group'] == group)]
        
        if user_row.empty:
            with data_lock:
                ordered_test_cases_df.loc[index, 'Actual Status'] = 'ERROR'
                ordered_test_cases_df.loc[index, 'Result'] = 'FAIL'
                ordered_test_cases_df.loc[index, 'Error Message'] = 'User not found'
            logger.info(f"[{current_thread}] User not found for Group: {group} in Environment: {selected_env}")
            continue

        user = user_row.iloc[0]['User']
        password = user_passwords[user]

        connection_key = (host_url, user)
        if connection_key in failed_connections:
            with data_lock:
                ordered_test_cases_df.loc[index, 'Actual Status'] = 'ERROR'
                ordered_test_cases_df.loc[index, 'Result'] = 'FAIL'
                ordered_test_cases_df.loc[index, 'Error Message'] = 'Previous connection attempt failed for this user and host URL'
            logger.info(f"[{current_thread}] Skipping Test Case Number {test_case_number} due to prior connection failure for Host URL: {host_url} and User: {user}")
            continue

        try:
            conn = get_or_create_trino_connection(host_url, user, password)
        except Exception as e:
            error_message = f"Connection failed: {str(e)}"
            with data_lock:
                ordered_test_cases_df.loc[index, 'Actual Status'] = 'ERROR'
                ordered_test_cases_df.loc[index, 'Result'] = 'FAIL'
                ordered_test_cases_df.loc[index, 'Error Message'] = error_message
            logger.info(f"[{current_thread}] Failed to connect for Test Case Number {test_case_number}: {error_message}")
            failed_connections.add(connection_key)
            continue
        
        actual_status, response = execute_sql_with_trino(conn, executed_sql, logger)
        with data_lock:
            ordered_test_cases_df.loc[index, 'Actual Status'] = actual_status
            ordered_test_cases_df.loc[index, 'Actual Response'] = str(response)

        if actual_status == 'ERROR':
            with data_lock:
                ordered_test_cases_df.loc[index, 'Error Message'] = response
        
        logger.info(f"[{current_thread}] Response for Test Case Number {test_case_number}: {response}")

        with data_lock:
            if actual_status == expected_status:
                # If statuses match, compare the expected and actual responses if provided
                if expected_response:
                    if str(response).strip() == expected_response:
                        ordered_test_cases_df.loc[index, 'Result'] = 'PASS'
                    else:
                        ordered_test_cases_df.loc[index, 'Result'] = 'FAIL'
                else:
                    # No Expected Response specified, consider it a pass
                    ordered_test_cases_df.loc[index, 'Result'] = 'PASS'
            else:
                ordered_test_cases_df.loc[index, 'Result'] = 'FAIL'

def process_test_cases(result_filepath, selected_env, selected_teams, user_passwords, ordered_test_cases_df, users_df, trino_env_df, sql_variables_df, refresh_frequency):
    global execution_complete  # Use the global execution_complete flag

    # Initialize columns
    ordered_test_cases_df['Actual Status'] = ""
    ordered_test_cases_df['Actual Response'] = ""
    ordered_test_cases_df['Result'] = ""
    ordered_test_cases_df['Executed SQL'] = ""
    ordered_test_cases_df['Error Message'] = ""

    if 'Expected Response' in ordered_test_cases_df.columns:
        ordered_test_cases_df['Expected Response'] = ordered_test_cases_df['Expected Response'].astype(str)
    else:
        ordered_test_cases_df['Expected Response'] = ""  # Initialize if not present
    
    total_test_cases = len(ordered_test_cases_df)

    # Collect all SQL variable values before execution
    collect_variable_values(ordered_test_cases_df, sql_variables_df, selected_env)

    # Start the summary display thread
    summary_thread = threading.Thread(target=display_summary, args=(ordered_test_cases_df, total_test_cases, refresh_frequency), daemon=True)
    summary_thread.start()

    # Execute setup test cases
    execute_test_cases(ordered_test_cases_df, trino_env_df, users_df, selected_env, user_passwords, sql_variables_df, "Setup")

    # Start threads for each team, ensuring only test cases for that team are executed
    team_threads = []
    for team in selected_teams:
        team_thread = threading.Thread(
            target=execute_test_cases, 
            args=(ordered_test_cases_df, trino_env_df, users_df, selected_env, user_passwords, sql_variables_df, "Test", team),
            name=f"Team-{team}"
        )
        team_threads.append(team_thread)
        team_thread.start()

    # Wait for all team threads to finish
    for thread in team_threads:
        thread.join()

    # Execute cleanup test cases
    execute_test_cases(ordered_test_cases_df, trino_env_df, users_df, selected_env, user_passwords, sql_variables_df, "Clean up")

    # Write the results to the Excel file first
    save_results_to_new_excel(result_filepath, ordered_test_cases_df)
    apply_result_formatting(result_filepath, ordered_test_cases_df)
    save_filtered_sheets(result_filepath, selected_env, users_df, trino_env_df, sql_variables_df)

    # Mark the execution as complete
    execution_complete = True
    # Wait for the summary thread to complete
    summary_thread.join()

if __name__ == '__main__':
    main()
